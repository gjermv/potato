# -*- coding: utf-8 -*-
"""
Created on Fri Oct 21 11:05:42 2022

@author: gjermund.vingerhagen
"""

import pandas as pd
import openpyxl

report_4b_nr = 19

def readPowerBI_xlsxfile(filetype='4b'):
    folder = "C:\\python_proj\\parseToSVVEcoReport\\inFiles\\"
    file_4b= "4.b Timer, uposterte og posterte ({})".format(report_4b_nr)
    filename =  folder + file_4b + ".xlsx"
    

    df = pd.read_excel(filename,skiprows=9)
    return df

def getXLSXFileName():
    folder = "C:\\python_proj\\parseToSVVEcoReport\\inFiles\\"
    file_4b= "4.b Timer, uposterte og posterte ({})".format(report_4b_nr)
    filename =  folder + file_4b + ".xlsx"
    
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    value1 = ws['C4'].value
    value2 = ws['C6'].value.replace('/','-')
    print(value2)
    return 'C:\\python_proj\\parseToSVVEcoReport\\outFiles\\' + '{} {}  -Export.xlsx'.format(value1,value2)

def b4_to_xlsxfile(inDF, filetype='4b'):

    myFilename =  getXLSXFileName()
    
    outColumns = ['Dato', 'Beskrivelse','Timer', 'Honorar', 'Aktivitet', 'Ansatt', 'Oppdragnr', 'Status']

    inDF.to_excel(myFilename, columns=outColumns )


def fix4B(df):
    df = df[~df['Ansattnr'].isna()]
    df = df.ffill()
    df = df.sort_values(by=['Dato'],ascending=True)
    df[['Oppdragnr','Oppdragnavn']] = df['Oppdrag'].str.split(' ', expand=True,n=1)
    return df

myDF =  readPowerBI_xlsxfile()
newDF = fix4B(myDF)
b4_to_xlsxfile(newDF)
